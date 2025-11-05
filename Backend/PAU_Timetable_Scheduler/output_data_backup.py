import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from collections import defaultdict
from datetime import datetime

class TimetableExporter:
    def __init__(self):
        self.saved_data_path = os.path.join(os.path.dirname(__file__), 'data', 'timetable_data.json')
        self.output_dir = os.path.join(os.path.dirname(__file__), 'output_data')
        
        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Load course, room, and faculty data
        self.course_data = self.load_course_data()
        self.room_data = self.load_room_data()
        self.faculty_data = self.load_faculty_data()
        
        # Keywords to identify SST (engineering) groups
        self.sst_keywords = [
            'engineering', 'eng', 'computer science', 'software engineering', 'data science',
            'mechatronics', 'electrical', 'mechanical', 'csc', 'sen', 'data', 'ds'
        ]
        
        # Time slots
        self.time_slots = [
            "9:00-9:50", "10:00-10:50", "11:00-11:50", "12:00-12:50", 
            "1:00-1:50", "2:00-2:50", "3:00-3:50", "4:00-4:50"
        ]
        
        # Days of the week
        self.days = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]

    def load_course_data(self):
        """Load course data from JSON file"""
        try:
            course_path = os.path.join(os.path.dirname(__file__), 'data', 'course-data.json')
            with open(course_path, 'r', encoding='utf-8') as f:
                courses = json.load(f)
            
            # Create a mapping from course code to course info
            course_map = {}
            for course in courses:
                course_map[course['code']] = course
            
            return course_map
        except Exception as e:
            print(f"❌ Error loading course data: {e}")
            return {}

    def load_room_data(self):
        """Load room data from JSON file"""
        try:
            room_path = os.path.join(os.path.dirname(__file__), 'data', 'rooms-data.json')
            with open(room_path, 'r', encoding='utf-8') as f:
                rooms = json.load(f)
            
            # Create a mapping from room name to room info
            room_map = {}
            for room in rooms:
                room_map[room['name']] = room
            
            return room_map
        except Exception as e:
            print(f"❌ Error loading room data: {e}")
            return {}

    def load_faculty_data(self):
        """Load faculty data from JSON file"""
        try:
            faculty_path = os.path.join(os.path.dirname(__file__), 'data', 'faculty-data.json')
            with open(faculty_path, 'r', encoding='utf-8') as f:
                faculty = json.load(f)
            
            # Create a mapping from faculty ID to faculty info
            faculty_map = {}
            for fac in faculty:
                faculty_map[fac['id']] = fac  # Use 'id' field instead of 'faculty_id'
            
            return faculty_map
        except Exception as e:
            print(f"❌ Error loading faculty data: {e}")
            return {}

    def load_saved_timetable_data(self):
        """Load the latest saved timetable data"""
        try:
            if os.path.exists(self.saved_data_path):
                with open(self.saved_data_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # FIX: Check for the new dictionary format containing 'timetables'
                if isinstance(data, dict) and 'timetables' in data:
                    print(f"✅ Loaded saved timetable data (new format): {len(data['timetables'])} groups")
                    return data['timetables'] # Return only the list of timetables
                else:
                    # Handle old format for backward compatibility
                    print(f"✅ Loaded saved timetable data (old format): {len(data)} groups")
                    return data
            else:
                print("❌ No saved timetable data found")
                return None
        except Exception as e:
            print(f"❌ Error loading saved data: {e}")
            return None

    def load_fresh_optimization_data(self):
        """Load fresh DE optimization data if available"""
        fresh_data_path = os.path.join(os.path.dirname(__file__), 'data', 'fresh_timetable_data.json')
        try:
            if os.path.exists(fresh_data_path):
                with open(fresh_data_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                print(f"✅ Loaded fresh optimization data: {len(data)} groups")
                return data
            else:
                print("❌ No fresh optimization data found")
                return None
        except Exception as e:
            print(f"❌ Error loading fresh optimization data: {e}")
            return None

    def get_timetable_data(self):
        """Get timetable data - prefer saved data, fallback to fresh optimization data"""
        # First try to load saved data (user modifications)
        data = self.load_saved_timetable_data()
        if data:
            return data
        
        # If no saved data, try to load fresh optimization data
        data = self.load_fresh_optimization_data()
        if data:
            return data
        
        return None

    def is_sst_group(self, group_name):
        """Check if a student group belongs to SST (engineering) based on keywords"""
        group_name_lower = group_name.lower()
        return any(keyword in group_name_lower for keyword in self.sst_keywords)

    def extract_main_program_name(self, group_name):
        """Extract the main program name from student group name"""
        # Remove year information and stream information
        # Examples: "Computer Science - Year 1" -> "Computer Science"
        #          "Electrical Engineering - Year 2/Stream 1" -> "Electrical Engineering"
        
        # First check for hyphen
        if ' - ' in group_name:
            main_name = group_name.split(' - ')[0]
        else:
            # If no hyphen, look for the word "year" (case insensitive)
            match = re.search(r'\s+year\s+\d+', group_name, re.IGNORECASE)
            if match:
                main_name = group_name[:match.start()]
            else:
                main_name = group_name
        
        # Remove any stream information that might be in the main name
        main_name = re.sub(r'/Stream \d+', '', main_name)
        main_name = re.sub(r'Stream \d+', '', main_name)
        
        return main_name.strip()

    def extract_lecturer_info(self, cell_content):
        """Extract lecturer information from cell content"""
        if not cell_content or cell_content in ["FREE", "BREAK", ""]:
            return None
        
        # Cell format: Course Code\nRoom Name\nFaculty
        lines = cell_content.split('\n')
        if len(lines) >= 3:
            return {
                'course_code': lines[0].strip(),
                'room': lines[1].strip(),
                'faculty': lines[2].strip()
            }
        return None

    def create_combined_program_sheet(self, wb, sheet_name, groups_data):
        """Create a combined timetable sheet for multiple student groups of the same program"""
        ws = wb.create_sheet(title=sheet_name)
        
        current_row = 1
        
        for group_idx, group_data in enumerate(groups_data):
            full_group_name = group_data['student_group']['name']
            
            # Add student group name header
            group_header_cell = ws.cell(row=current_row, column=1, value=full_group_name)
            group_header_cell.font = Font(bold=True, size=14)
            group_header_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Create headers - Fixed order: Course Code, Course Name, Units, then TIME, DAY, CLASSROOM, BUILDING pattern
            headers = [
                ("Course Code", "A"),
                ("Course Name", "B"), 
                ("Units", "C"),
                ("TIME", "D"),
                ("MONDAY", "E"),
                ("CLASSROOM", "F"),
                ("BUILDING", "G"),
                ("TIME", "H"),
                ("TUESDAY", "I"),
                ("CLASSROOM", "J"),
                ("BUILDING", "K"),
                ("TIME", "L"),
                ("WEDNESDAY", "M"),
                ("CLASSROOM", "N"),
                ("BUILDING", "O"),
                ("TIME", "P"),
                ("THURSDAY", "Q"),
                ("CLASSROOM", "R"),
                ("BUILDING", "S"),
                ("TIME", "T"),
                ("FRIDAY", "U"),
                ("CLASSROOM", "V"),
                ("BUILDING", "W")
            ]
            
            # Set headers and apply formatting
            for col_idx, (header_text, _) in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header_text)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply green color to Time and day columns
                if header_text in ["TIME", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]:
                    cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                
                # Apply borders to headers from TIME column onwards (column 4+)
                if col_idx >= 4:  # TIME column and onwards
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
            
            current_row += 1
            
            # Extract unique courses from timetable
            courses_data = self.extract_courses_from_timetable(group_data['timetable'])
            
            # Fill time slots and schedule data
            for time_idx in range(len(self.time_slots)):
                time_row = current_row + time_idx
                
                # Define column groups for each day (TIME, DAY, CLASSROOM, BUILDING)
                day_column_groups = [
                    (4, 5, 6, 7),   # Monday: D, E, F, G
                    (8, 9, 10, 11), # Tuesday: H, I, J, K
                    (12, 13, 14, 15), # Wednesday: L, M, N, O
                    (16, 17, 18, 19), # Thursday: P, Q, R, S
                    (20, 21, 22, 23)  # Friday: T, U, V, W
                ]
                
                for day_idx, (time_col, day_col, classroom_col, building_col) in enumerate(day_column_groups):
                    # Time column
                    time_cell = ws.cell(row=time_row, column=time_col, value=self.time_slots[time_idx])
                    time_cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                    time_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Check if this is lunch break time (1:00-1:50) and specific days
                    is_lunch_break = (time_idx == 4 and day_idx in [0, 2, 4])  # Monday(0), Wednesday(2), Friday(4)
                    
                    if is_lunch_break:
                        # Day column (BREAK)
                        day_cell = ws.cell(row=time_row, column=day_col, value="BREAK")
                        day_cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                        day_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Empty classroom and building for break
                        classroom_cell = ws.cell(row=time_row, column=classroom_col, value="")
                        classroom_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        building_cell = ws.cell(row=time_row, column=building_col, value="")
                        building_cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        # Get class info for this time and day
                        class_info = self.get_class_at_time_day(group_data['timetable'], time_idx, day_idx)
                        
                        # Day column (course code)
                        day_cell = ws.cell(row=time_row, column=day_col, value=class_info['course_code'] if class_info else "")
                        day_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Classroom column
                        classroom_cell = ws.cell(row=time_row, column=classroom_col, value=class_info['room'] if class_info else "")
                        classroom_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Building column
                        building = self.get_building_from_room(class_info['room']) if class_info else ""
                        building_cell = ws.cell(row=time_row, column=building_col, value=building)
                        building_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Apply borders to all timetable cells (from TIME column onwards)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    time_cell.border = border
                    day_cell.border = border
                    classroom_cell.border = border
                    building_cell.border = border
            
            # Fill course information (A, B, C columns) - NO BORDERS for these
            course_row = current_row
            for course_code, course_info in courses_data.items():
                # Course Code (A)
                code_cell = ws.cell(row=course_row, column=1, value=course_code)
                code_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Course Name (B)
                course_name = self.get_course_name(course_code)
                name_cell = ws.cell(row=course_row, column=2, value=course_name)
                name_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Units (C)
                units = course_info.get('hours', 1)
                units_cell = ws.cell(row=course_row, column=3, value=units)
                units_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                course_row += 1
            
            # Move to next student group (4 lines below)
            current_row += len(self.time_slots) + 4
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column].width = adjusted_width

    def create_student_group_timetable_sheet(self, wb, sheet_name, group_data):
        """Create a timetable sheet for a student group"""
        ws = wb.create_sheet(title=sheet_name)
        
        # Create header row with TIME and days
        headers = ["TIME"] + self.days
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Fill in the timetable data
        timetable_rows = group_data['timetable']
        
        for row_idx, row_data in enumerate(timetable_rows):
            excel_row = row_idx + 2  # Start from row 2 (after header)
            
            # Add time slot
            time_cell = ws.cell(row=excel_row, column=1, value=self.time_slots[row_idx])
            time_cell.font = Font(bold=True, color="FFFFFF")
            time_cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
            time_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add class data for each day
            for day_idx in range(len(self.days)):
                col = day_idx + 2  # Start from column 2 (after TIME column)
                
                if day_idx + 1 < len(row_data):  # Skip time column in source data
                    cell_content = row_data[day_idx + 1]
                    
                    if cell_content and cell_content not in ["FREE", ""]:
                        # Parse cell content
                        lines = cell_content.split('\n')
                        if len(lines) >= 3:
                            course_code = lines[0]
                            room = lines[1]
                            faculty = lines[2]
                            
                            # Format as: Course Code, Room, Faculty
                            display_text = f"{course_code}\n{room}\n{faculty}"
                        else:
                            display_text = cell_content
                    else:
                        display_text = ""
                    
                    cell = ws.cell(row=excel_row, column=col, value=display_text)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Apply border
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column].width = adjusted_width
        
        # Set row heights for better visibility
        for row in range(1, len(timetable_rows) + 2):
            ws.row_dimensions[row].height = 60

    def create_combined_lecturer_sheet(self, wb, sheet_name, lecturers_data):
        """Create a combined timetable sheet for all lecturers"""
        ws = wb.create_sheet(title=sheet_name)
        
        current_row = 1
        
        for lecturer_name, lecturer_schedule in lecturers_data.items():
            # Add lecturer name header
            lecturer_header_cell = ws.cell(row=current_row, column=1, value=lecturer_name)
            lecturer_header_cell.font = Font(bold=True, size=14)
            lecturer_header_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Create headers - Fixed order: Course Code, Course Name, Units, then TIME, DAY, CLASSROOM, BUILDING pattern
            headers = [
                ("Course Code", "A"),
                ("Course Name", "B"), 
                ("Units", "C"),
                ("TIME", "D"),
                ("MONDAY", "E"),
                ("CLASSROOM", "F"),
                ("BUILDING", "G"),
                ("TIME", "H"),
                ("TUESDAY", "I"),
                ("CLASSROOM", "J"),
                ("BUILDING", "K"),
                ("TIME", "L"),
                ("WEDNESDAY", "M"),
                ("CLASSROOM", "N"),
                ("BUILDING", "O"),
                ("TIME", "P"),
                ("THURSDAY", "Q"),
                ("CLASSROOM", "R"),
                ("BUILDING", "S"),
                ("TIME", "T"),
                ("FRIDAY", "U"),
                ("CLASSROOM", "V"),
                ("BUILDING", "W")
            ]
            
            # Set headers and apply formatting
            for col_idx, (header_text, _) in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header_text)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply green color to Time and day columns
                if header_text in ["TIME", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]:
                    cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                
                # Apply borders to headers from TIME column onwards (column 4+)
                if col_idx >= 4:  # TIME column and onwards
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
            
            current_row += 1
            
            # Extract unique courses from lecturer schedule
            courses_data = self.extract_courses_from_lecturer_schedule(lecturer_schedule)
            
            # Create schedule grid
            schedule_grid = {}
            for time_idx in range(len(self.time_slots)):
                schedule_grid[time_idx] = {}
                for day_idx in range(len(self.days)):
                    schedule_grid[time_idx][day_idx] = None
            
            # Fill lecturer's schedule
            for class_info in lecturer_schedule:
                time_idx = class_info['time_slot']
                day_idx = class_info['day']
                schedule_grid[time_idx][day_idx] = class_info
            
            # Fill time slots and schedule data
            for time_idx in range(len(self.time_slots)):
                time_row = current_row + time_idx
                
                # Define column groups for each day (TIME, DAY, CLASSROOM, BUILDING)
                day_column_groups = [
                    (4, 5, 6, 7),   # Monday: D, E, F, G
                    (8, 9, 10, 11), # Tuesday: H, I, J, K
                    (12, 13, 14, 15), # Wednesday: L, M, N, O
                    (16, 17, 18, 19), # Thursday: P, Q, R, S
                    (20, 21, 22, 23)  # Friday: T, U, V, W
                ]
                
                for day_idx, (time_col, day_col, classroom_col, building_col) in enumerate(day_column_groups):
                    # Time column
                    time_cell = ws.cell(row=time_row, column=time_col, value=self.time_slots[time_idx])
                    time_cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                    time_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Check if this is lunch break time (1:00-1:50) and specific days
                    is_lunch_break = (time_idx == 4 and day_idx in [0, 2, 4])  # Monday(0), Wednesday(2), Friday(4)
                    
                    if is_lunch_break:
                        # Day column (BREAK)
                        day_cell = ws.cell(row=time_row, column=day_col, value="BREAK")
                        day_cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                        day_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Empty classroom and building for break
                        classroom_cell = ws.cell(row=time_row, column=classroom_col, value="")
                        classroom_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        building_cell = ws.cell(row=time_row, column=building_col, value="")
                        building_cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        # Get class info for this time and day
                        class_info = schedule_grid[time_idx][day_idx]
                        
                        # Day column (course code)
                        day_cell = ws.cell(row=time_row, column=day_col, value=class_info['course_code'] if class_info else "")
                        day_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Classroom column
                        classroom_cell = ws.cell(row=time_row, column=classroom_col, value=class_info['room'] if class_info else "")
                        classroom_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Building column
                        building = self.get_building_from_room(class_info['room']) if class_info else ""
                        building_cell = ws.cell(row=time_row, column=building_col, value=building)
                        building_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Apply borders to all timetable cells (from TIME column onwards)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    time_cell.border = border
                    day_cell.border = border
                    classroom_cell.border = border
                    building_cell.border = border
            
            # Fill course information (A, B, C columns) - NO BORDERS for these
            course_row = current_row
            for course_code, course_info in courses_data.items():
                # Course Code (A)
                code_cell = ws.cell(row=course_row, column=1, value=course_code)
                code_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Course Name (B)
                course_name = self.get_course_name(course_code)
                name_cell = ws.cell(row=course_row, column=2, value=course_name)
                name_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Units (C)
                units = course_info.get('hours', 1)
                units_cell = ws.cell(row=course_row, column=3, value=units)
                units_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                course_row += 1
            
            # Move to next lecturer (4 lines below)
            current_row += len(self.time_slots) + 4
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column].width = adjusted_width

    def export_sst_timetables(self):
        """Export SST (engineering) timetables"""
        data = self.get_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Group SST student groups by main program
            sst_programs = defaultdict(list)
            
            for group_data in data:
                group_name = group_data['student_group']['name']
                if self.is_sst_group(group_name):
                    main_program = self.extract_main_program_name(group_name)
                    sst_programs[main_program].append(group_data)
            
            if not sst_programs:
                return False, "No SST student groups found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets for each unique program (one sheet per main program)
            for program_name, groups in sst_programs.items():
                # Sort groups by year for better organization
                groups.sort(key=lambda x: x['student_group']['name'])
                
                # Use main program name as sheet name, but limit length for Excel
                safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
                self.create_combined_program_sheet(wb, safe_sheet_name, groups)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SST_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"SST Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting SST timetables: {str(e)}"

    def export_tyd_timetables(self):
        """Export TYD (non-engineering) timetables"""
        data = self.get_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Group TYD student groups by main program
            tyd_programs = defaultdict(list)
            
            for group_data in data:
                group_name = group_data['student_group']['name']
                if not self.is_sst_group(group_name):  # Not SST = TYD
                    main_program = self.extract_main_program_name(group_name)
                    tyd_programs[main_program].append(group_data)
            
            if not tyd_programs:
                return False, "No TYD student groups found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets for each unique program (one sheet per main program)
            for program_name, groups in tyd_programs.items():
                # Sort groups by year for better organization
                groups.sort(key=lambda x: x['student_group']['name'])
                
                # Use main program name as sheet name, but limit length for Excel
                safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
                self.create_combined_program_sheet(wb, safe_sheet_name, groups)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"TYD_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"TYD Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting TYD timetables: {str(e)}"

    def export_lecturer_timetables(self):
        """Export all lecturer timetables"""
        data = self.get_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Collect all lecturer schedules
            lecturer_schedules = defaultdict(list)
            
            for group_data in data:
                student_group_name = group_data['student_group']['name']
                timetable_rows = group_data['timetable']
                
                for time_slot_idx, row_data in enumerate(timetable_rows):
                    for day_idx in range(len(self.days)):
                        if day_idx + 1 < len(row_data):  # Skip time column
                            cell_content = row_data[day_idx + 1]
                            lecturer_info = self.extract_lecturer_info(cell_content)
                            
                            if lecturer_info and lecturer_info['course_code']:
                                course_code = lecturer_info['course_code']
                                
                                # Get all faculty for this course from course data
                                course_info = self.course_data.get(course_code, {})
                                faculty_ids = course_info.get('facultyId', [])
                                
                                # If facultyId is a string, convert to list
                                if isinstance(faculty_ids, str):
                                    faculty_ids = [faculty_ids]
                                
                                # Create schedule entries for ALL faculty members
                                for faculty_id in faculty_ids:
                                    # Get faculty name from faculty data
                                    faculty_info = self.faculty_data.get(faculty_id, {})
                                    lecturer_name = faculty_info.get('name', faculty_id)
                                    
                                    # Skip generic entries
                                    if lecturer_name.lower() not in ['unknown', 'tbd', 'staff', '']:
                                        lecturer_schedules[lecturer_name].append({
                                            'time_slot': time_slot_idx,
                                            'day': day_idx,
                                            'course_code': course_code,
                                            'room': lecturer_info['room'],
                                            'student_group': student_group_name
                                        })
            
            if not lecturer_schedules:
                return False, "No lecturer data found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create one combined sheet for all lecturers
            self.create_combined_lecturer_sheet(wb, "Lecturer Timetables", lecturer_schedules)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Lecturer_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"Lecturer Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting lecturer timetables: {str(e)}"

    def extract_courses_from_timetable(self, timetable_rows):
        """Extract unique courses from timetable data"""
        courses = {}
        
        for row_idx, row_data in enumerate(timetable_rows):
            for day_idx in range(len(self.days)):
                if day_idx + 1 < len(row_data):  # Skip time column
                    cell_content = row_data[day_idx + 1]
                    
                    if cell_content and cell_content not in ["FREE", "", "BREAK"]:
                        lines = cell_content.split('\n')
                        if len(lines) >= 1:
                            course_code = lines[0].strip()
                            if course_code:
                                if course_code not in courses:
                                    courses[course_code] = {'hours': 0}
                                courses[course_code]['hours'] += 1
        
        return courses

    def extract_courses_from_lecturer_schedule(self, lecturer_schedule):
        """Extract unique courses from lecturer schedule data"""
        courses = {}
        
        for class_info in lecturer_schedule:
            course_code = class_info['course_code']
            if course_code:
                if course_code not in courses:
                    courses[course_code] = {'hours': 0}
                courses[course_code]['hours'] += 1
        
        return courses

    def get_class_at_time_day(self, timetable_rows, time_idx, day_idx):
        """Get class information at specific time and day"""
        if time_idx < len(timetable_rows):
            row_data = timetable_rows[time_idx]
            if day_idx + 1 < len(row_data):  # Skip time column
                cell_content = row_data[day_idx + 1]
                
                if cell_content and cell_content not in ["FREE", "", "BREAK"]:
                    lines = cell_content.split('\n')
                    if len(lines) >= 3:
                        return {
                            'course_code': lines[0].strip(),
                            'room': lines[1].strip(),
                            'faculty': lines[2].strip()
                        }
        return None

    def get_building_from_room(self, room_name):
        """Determine building from room data"""
        if not room_name:
            return ""
        
        # Look up room in the loaded room data
        room_info = self.room_data.get(room_name)
        if room_info and 'building' in room_info:
            return room_info['building']
        
        # Fallback to heuristic if not found in data
        room_lower = room_name.lower()
        if any(keyword in room_lower for keyword in ['eng', 'lab', 'workshop', 'tech']):
            return "SST"
        else:
            return "TYD"

    def get_course_name(self, course_code):
        """Get full course name from course code using loaded course data"""
        if not course_code:
            return ""
        
        # Look up course in the loaded course data
        course_info = self.course_data.get(course_code)
        if course_info and 'name' in course_info:
            return course_info['name']
        
        # Fallback if not found in data
        return f"{course_code} (Course Name)"

# Main functions for easy import
def export_sst_timetables():
    """Export SST timetables"""
    exporter = TimetableExporter()
    return exporter.export_sst_timetables()

def export_tyd_timetables():
    """Export TYD timetables"""
    exporter = TimetableExporter()
    return exporter.export_tyd_timetables()

def export_lecturer_timetables():
    """Export lecturer timetables"""
    exporter = TimetableExporter()
    return exporter.export_lecturer_timetables()

# New methods that return bytes for browser downloads
def export_sst_timetables_bytes():
    """Export SST timetables and return bytes for download"""
    exporter = TimetableExporter()
    data = exporter.get_timetable_data()
    if not data:
        return None, "No timetable data available"
    
    try:
        # Group SST student groups by main program
        sst_programs = defaultdict(list)
        
        for group_data in data:
            group_name = group_data['student_group']['name']
            if exporter.is_sst_group(group_name):
                main_program = exporter.extract_main_program_name(group_name)
                sst_programs[main_program].append(group_data)
        
        if not sst_programs:
            return None, "No SST student groups found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for program_name, groups in sst_programs.items():
            groups.sort(key=lambda x: x['student_group']['name'])
            safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
            exporter.create_combined_program_sheet(wb, safe_sheet_name, groups)
        
        # Save to BytesIO
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SST_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        return None, f"Error exporting SST timetables: {str(e)}"

def export_tyd_timetables_bytes():
    """Export TYD timetables and return bytes for download"""
    exporter = TimetableExporter()
    data = exporter.get_timetable_data()
    if not data:
        return None, "No timetable data available"
    
    try:
        # Group TYD student groups by main program
        tyd_programs = defaultdict(list)
        
        for group_data in data:
            group_name = group_data['student_group']['name']
            if not exporter.is_sst_group(group_name):
                main_program = exporter.extract_main_program_name(group_name)
                tyd_programs[main_program].append(group_data)
        
        if not tyd_programs:
            return None, "No TYD student groups found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for program_name, groups in tyd_programs.items():
            groups.sort(key=lambda x: x['student_group']['name'])
            safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
            exporter.create_combined_program_sheet(wb, safe_sheet_name, groups)
        
        # Save to BytesIO
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TYD_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        return None, f"Error exporting TYD timetables: {str(e)}"

def export_lecturer_timetables_bytes():
    """Export lecturer timetables and return bytes for download"""
    exporter = TimetableExporter()
    data = exporter.get_timetable_data()
    if not data:
        return None, "No timetable data available"
    
    try:
        # Aggregate lecturer schedules
        lecturer_schedules = defaultdict(lambda: {
            'name': '',
            'courses': set(),
            'schedule': [[None for _ in range(5)] for _ in range(9)]
        })
        
        for group_data in data:
            group_name = group_data['student_group']['name']
            timetable = group_data.get('timetable', [])
            
            for day_idx, day_schedule in enumerate(timetable):
                for slot_idx, slot in enumerate(day_schedule):
                    if slot and slot != 'Free':
                        try:
                            parts = slot.split('\n')
                            course_part = parts[0].strip()
                            lecturer_part = parts[1].strip() if len(parts) > 1 else ''
                            
                            match = re.search(r'\((.*?)\)', lecturer_part)
                            if match:
                                lecturer_id = match.group(1).strip()
                                lecturer_name = exporter.get_lecturer_name(lecturer_id)
                                
                                if lecturer_id not in lecturer_schedules:
                                    lecturer_schedules[lecturer_id]['name'] = lecturer_name
                                
                                lecturer_schedules[lecturer_id]['courses'].add(course_part)
                                
                                existing = lecturer_schedules[lecturer_id]['schedule'][slot_idx][day_idx]
                                if existing:
                                    if group_name not in existing:
                                        lecturer_schedules[lecturer_id]['schedule'][slot_idx][day_idx] = f"{existing}, {group_name}"
                                else:
                                    lecturer_schedules[lecturer_id]['schedule'][slot_idx][day_idx] = f"{course_part} - {group_name}"
                        except Exception:
                            continue
        
        if not lecturer_schedules:
            return None, "No lecturer schedules found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for lecturer_id, data in sorted(lecturer_schedules.items()):
            lecturer_name = data['name'] or lecturer_id
            safe_sheet_name = re.sub(r'[^\w\s-]', '', lecturer_name)[:31]
            exporter.create_lecturer_sheet(wb, safe_sheet_name, data['schedule'])
        
        # Save to BytesIO
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Lecturer_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        return None, f"Error exporting lecturer timetables: {str(e)}"

# New functions that accept timetable data directly from Dash UI
def export_sst_timetables_bytes_from_data(timetable_data):
    """Export SST timetables from provided data and return bytes for download"""
    if not timetable_data:
        return None, "No timetable data available"
    
    exporter = TimetableExporter()
    
    # Extract timetables list from the data structure
    data = timetable_data.get('timetables', []) if isinstance(timetable_data, dict) else timetable_data
    
    if not data:
        return None, "No timetable data available"
    
    try:
        # Group SST student groups by main program
        sst_programs = defaultdict(list)
        
        for group_data in data:
            group_name = group_data['student_group']['name']
            if exporter.is_sst_group(group_name):
                main_program = exporter.extract_main_program_name(group_name)
                sst_programs[main_program].append(group_data)
        
        if not sst_programs:
            return None, "No SST student groups found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for program_name, groups in sst_programs.items():
            groups.sort(key=lambda x: x['student_group']['name'])
            safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
            exporter.create_combined_program_sheet(wb, safe_sheet_name, groups)
        
        # Save to BytesIO
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SST_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        return None, f"Error exporting SST timetables: {str(e)}"

def export_tyd_timetables_bytes_from_data(timetable_data):
    """Export TYD timetables from provided data and return bytes for download"""
    if not timetable_data:
        return None, "No timetable data available"
    
    exporter = TimetableExporter()
    
    # Extract timetables list from the data structure
    data = timetable_data.get('timetables', []) if isinstance(timetable_data, dict) else timetable_data
    
    if not data:
        return None, "No timetable data available"
    
    try:
        # Group TYD student groups by main program
        tyd_programs = defaultdict(list)
        
        for group_data in data:
            group_name = group_data['student_group']['name']
            if not exporter.is_sst_group(group_name):
                main_program = exporter.extract_main_program_name(group_name)
                tyd_programs[main_program].append(group_data)
        
        if not tyd_programs:
            return None, "No TYD student groups found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for program_name, groups in tyd_programs.items():
            groups.sort(key=lambda x: x['student_group']['name'])
            safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
            exporter.create_combined_program_sheet(wb, safe_sheet_name, groups)
        
        # Save to BytesIO
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TYD_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        return None, f"Error exporting TYD timetables: {str(e)}"

def export_lecturer_timetables_bytes_from_data(timetable_data):
    """Export lecturer timetables from provided data and return bytes for download"""
    if not timetable_data:
        return None, "No timetable data available"
    
    exporter = TimetableExporter()
    
    # Extract timetables list from the data structure
    data = timetable_data.get('timetables', []) if isinstance(timetable_data, dict) else timetable_data
    
    if not data:
        return None, "No timetable data available"
    
    try:
        # Aggregate lecturer schedules
        lecturer_schedules = defaultdict(lambda: {
            'name': '',
            'courses': set(),
            'schedule': [[None for _ in range(5)] for _ in range(9)]
        })
        
        for group_data in data:
            group_name = group_data['student_group']['name']
            timetable = group_data.get('timetable', [])
            
            for day_idx, day_schedule in enumerate(timetable):
                for slot_idx, slot in enumerate(day_schedule):
                    if slot and slot != 'Free' and slot != 'BREAK':
                        try:
                            # Parse the cell content
                            lines = slot.split('\n')
                            if not lines:
                                continue
                            
                            # Extract course info
                            course_part = lines[0].strip()
                            
                            # Extract lecturer info
                            lecturer_id = None
                            lecturer_name = None
                            
                            for line in lines:
                                if 'Lecturer:' in line:
                                    # Format: "Lecturer: Name (ID)" or "Lecturer: Name"
                                    lecturer_part = line.replace('Lecturer:', '').strip()
                                    match = re.search(r'\((.*?)\)', lecturer_part)
                                    if match:
                                        lecturer_id = match.group(1).strip()
                                        lecturer_name = lecturer_part.split('(')[0].strip()
                                    else:
                                        lecturer_name = lecturer_part
                                        lecturer_id = lecturer_part
                                    break
                            
                            if not lecturer_id:
                                continue
                            
                            if lecturer_id not in lecturer_schedules:
                                lecturer_schedules[lecturer_id]['name'] = lecturer_name or lecturer_id
                            
                            lecturer_schedules[lecturer_id]['courses'].add(course_part)
                            
                            # Build schedule entry
                            existing = lecturer_schedules[lecturer_id]['schedule'][slot_idx][day_idx]
                            if existing:
                                if group_name not in existing:
                                    lecturer_schedules[lecturer_id]['schedule'][slot_idx][day_idx] = f"{existing}, {group_name}"
                            else:
                                lecturer_schedules[lecturer_id]['schedule'][slot_idx][day_idx] = f"{course_part} - {group_name}"
                        except Exception as e:
                            continue
        
        if not lecturer_schedules:
            return None, "No lecturer schedules found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for lecturer_id, lec_data in sorted(lecturer_schedules.items()):
            lecturer_name = lec_data['name'] or lecturer_id
            safe_sheet_name = re.sub(r'[^\w\s-]', '', lecturer_name)[:31]
            exporter.create_lecturer_sheet(wb, safe_sheet_name, lec_data['schedule'])
        
        # Save to BytesIO
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Lecturer_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        return None, f"Error exporting lecturer timetables: {str(e)}"

if __name__ == "__main__":
    # Test the export functions
    print("Testing SST export...")
    success, message = export_sst_timetables()
    print(f"SST: {message}")
    
    print("\nTesting TYD export...")
    success, message = export_tyd_timetables()
    print(f"TYD: {message}")
    
    print("\nTesting Lecturer export...")
    success, message = export_lecturer_timetables()
    print(f"Lecturer: {message}")
